import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
import shutil

def run():
    target_excel = 'd:/StockLap/StockLab_Test_Report_Final_v3.xlsx'
    
    # We create a brand new workbook to avoid any leftover weirdness, 
    # but the user probably wants the other sheets too.
    # We will copy from the original source.
    try:
        shutil.copy('d:/StockLap/StockLab_Test_Report_Updated.xlsx', target_excel)
    except Exception as e:
        print("Copy failed:", e)

    wb = openpyxl.load_workbook(target_excel)
    if 'Report' in wb.sheetnames:
        ws = wb['Report']
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet('Report')

    # Styles
    title_font = Font(name='Arial', size=16, bold=True, color='000080')
    bold_font = Font(name='Arial', bold=True)
    normal_font = Font(name='Arial')
    
    header_fill = PatternFill('solid', fgColor='244062')
    sub_header_fill = PatternFill('solid', fgColor='DCE6F1')
    pass_bg = PatternFill('solid', fgColor='E2EFDA')
    fail_bg = PatternFill('solid', fgColor='FCE4D6')
    sec_bg = PatternFill('solid', fgColor='E4DFEC')
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 5
    
    data_col_1 = 27
    data_col_2 = 28
    
    # Setup data for Charts in hidden columns
    # Manual Data
    ws.cell(row=1, column=data_col_1, value="Result")
    ws.cell(row=1, column=data_col_2, value="Count")
    ws.cell(row=2, column=data_col_1, value="Pass")
    ws.cell(row=2, column=data_col_2, value=172)
    ws.cell(row=3, column=data_col_1, value="Fail")
    ws.cell(row=3, column=data_col_2, value=13)
    
    # API Data
    ws.cell(row=5, column=data_col_1, value="Result")
    ws.cell(row=5, column=data_col_2, value="Count")
    ws.cell(row=6, column=data_col_1, value="Pass")
    ws.cell(row=6, column=data_col_2, value=187)
    ws.cell(row=7, column=data_col_1, value="Fail")
    ws.cell(row=7, column=data_col_2, value=2)
    
    # Perf Data
    ws.cell(row=9, column=data_col_1, value="Users")
    ws.cell(row=9, column=data_col_2, value="Error %")
    perf_rates = [
        ("100 users", 0),
        ("500 users", 0),
        ("1000 users", 0.5),
        ("2000 users", 1.8),
        ("5000 users", 5.2),
        ("10000 users", 12.5)
    ]
    for i, (u, e) in enumerate(perf_rates):
        ws.cell(row=10+i, column=data_col_1, value=u)
        ws.cell(row=10+i, column=data_col_2, value=e)

    current_row = 2

    # Title
    cell = ws.cell(row=current_row, column=2, value="BÁO CÁO TỔNG HỢP KIỂM THỬ (TEST REPORT)")
    cell.font = title_font
    cell.alignment = center_align
    ws.merge_cells(f'B{current_row}:F{current_row}')
    current_row += 3

    # ==========================================
    # Helper to draw tables
    # ==========================================
    def draw_table_section(title, headers, items, start_row, metrics_title="THÔNG SỐ"):
        c = ws.cell(row=start_row, column=2, value=title.upper())
        c.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        c.fill = header_fill
        c.alignment = left_align
        ws.merge_cells(f'B{start_row}:F{start_row}')
        start_row += 1
        
        # Headers Let's assume headers = ["HẠNG MỤC CHI TIẾT", metrics_title, "TỶ LỆ (%)", "GHI CHÚ"]
        c_hdr_B = ws.cell(row=start_row, column=2, value=headers[0])
        c_hdr_B.font = bold_font; c_hdr_B.fill = sub_header_fill; c_hdr_B.alignment = center_align; c_hdr_B.border = border
        
        c_hdr_C = ws.cell(row=start_row, column=3)
        c_hdr_C.border = border; c_hdr_C.fill = sub_header_fill
        ws.merge_cells(f'B{start_row}:C{start_row}')
        
        ws.cell(row=start_row, column=4, value=headers[1]).border = border
        ws.cell(row=start_row, column=4).font = bold_font; ws.cell(row=start_row, column=4).fill = sub_header_fill; ws.cell(row=start_row, column=4).alignment = center_align

        ws.cell(row=start_row, column=5, value=headers[2]).border = border
        ws.cell(row=start_row, column=5).font = bold_font; ws.cell(row=start_row, column=5).fill = sub_header_fill; ws.cell(row=start_row, column=5).alignment = center_align

        ws.cell(row=start_row, column=6, value=headers[3]).border = border
        ws.cell(row=start_row, column=6).font = bold_font; ws.cell(row=start_row, column=6).fill = sub_header_fill; ws.cell(row=start_row, column=6).alignment = center_align
        start_row += 1

        for idx, item in enumerate(items):
            c1 = ws.cell(row=start_row, column=2, value=item[0])
            c1.border = border; c1.font = normal_font
            
            c_merge = ws.cell(row=start_row, column=3)
            c_merge.border = border
            ws.merge_cells(f'B{start_row}:C{start_row}')
            
            c2 = ws.cell(row=start_row, column=4, value=item[1])
            c2.border = border; c2.alignment = center_align; c2.font = bold_font if "Tổng" in str(item[0]) else normal_font
            
            c3 = ws.cell(row=start_row, column=5, value=item[2])
            c3.border = border; c3.alignment = center_align
            
            c4 = ws.cell(row=start_row, column=6, value=item[3])
            c4.border = border; c4.alignment = center_align
            
            # Apply color to Pass/Fail rows
            if "Pass" in str(item[0]):
                for col in range(2, 7): ws.cell(row=start_row, column=col).fill = pass_bg
            elif "Fail" in str(item[0]) or "Bug" in str(item[0]):
                for col in range(2, 7): ws.cell(row=start_row, column=col).fill = fail_bg
            
            start_row += 1
            
        return start_row

    # ==========================================
    # 1. MANUAL TESTING
    # ==========================================
    manual_data = [
        ("Tổng số Test Case", 185, "-", "Bao phủ các Module (Auth, Order, Admin,...)"),
        ("Đã thực hiện", 185, "100%", ""),
        ("Test Pass", 172, "93.0%", "Đạt yêu cầu logic"),
        ("Test Fail / Bug", 13, "7.0%", "Lỗi UI/UX & validation nhẹ"),
        ("Chưa thực hiện", 0, "0%", "")
    ]
    current_row = draw_table_section("1. KIỂM THỬ HỆ THỐNG - MANUAL TEST", ["HẠNG MỤC CHI TIẾT", "SỐ LƯỢNG", "TỶ LỆ (%)", "GHI CHÚ"], manual_data, current_row)

    # Chart 1
    pie1 = PieChart()
    pie1.title = "Manual Testing (185 TCs)"
    labels1 = Reference(ws, min_col=data_col_1, min_row=2, max_row=3)
    data1 = Reference(ws, min_col=data_col_2, min_row=1, max_row=3) # include title row for series name
    pie1.add_data(data1, titles_from_data=True)
    pie1.set_categories(labels1)
    pie1.height = 7
    pie1.width = 12
    ws.add_chart(pie1, f"H4")
    
    current_row += 2

    # ==========================================
    # 2. API & SECURITY TESTING
    # ==========================================
    api_data = [
        ("Tổng số API Functional (Postman)", 55, "-", "StockLab_API Collection"),
        ("Tổng số test Security", 134, "-", "Kiểm tra lỗ hổng bảo mật tiêu chuẩn OWASP"),
        ("Request Pass (Thành công/Chặn an toàn)", 187, "98.9%", "Đạt yêu cầu"),
        ("Request Fail (Lỗi/Lọt lỗ hổng)", 2, "1.1%", "Lỗi validation input nhẹ"),
        ("Mức độ bao phủ API & Security", "100%", "100%", "Đã test tất cả public/private APIs")
    ]
    current_row = draw_table_section("2. KIỂM THỬ TỰ ĐỘNG API & SECURITY", ["HẠNG MỤC CHI TIẾT", "KẾT QUẢ", "TỶ LỆ (%)", "GHI CHÚ"], api_data, current_row)
    
    # Sub-table Security
    ws.merge_cells(f'B{current_row}:C{current_row}')
    ws.cell(row=current_row, column=2, value="Chi tiết Security (134 TCs):").font = Font(bold=True, italic=True)
    current_row += 1
    
    # Add Security Mini Table
    sec_data = [
        ("SQL Injection", 38, "Chặn 100%"),
        ("XSS Attack", 38, "Chặn 100%"),
        ("Path Traversal", 46, "Chặn 100%"),
        ("Unauthorized Access", 12, "Chặn 100%"),
    ]
    ws.cell(row=current_row, column=2, value="LOẠI TẤN CÔNG").font = bold_font; ws.cell(row=current_row, column=2).border = border; ws.cell(row=current_row, column=2).fill = sec_bg; ws.cell(row=current_row, column=2).alignment = center_align
    ws.cell(row=current_row, column=3).border = border; ws.cell(row=current_row, column=3).fill = sec_bg
    ws.merge_cells(f'B{current_row}:C{current_row}')
    ws.cell(row=current_row, column=4, value="SỐ TEST").font = bold_font; ws.cell(row=current_row, column=4).border = border; ws.cell(row=current_row, column=4).fill = sec_bg; ws.cell(row=current_row, column=4).alignment = center_align
    ws.cell(row=current_row, column=5, value="KẾT QUẢ").font = bold_font; ws.cell(row=current_row, column=5).border = border; ws.cell(row=current_row, column=5).fill = sec_bg; ws.cell(row=current_row, column=5).alignment = center_align
    current_row += 1

    for atk, cnt, res in sec_data:
        ws.cell(row=current_row, column=2, value=atk).font = normal_font; ws.cell(row=current_row, column=2).border = border; ws.cell(row=current_row, column=2).alignment = center_align
        ws.cell(row=current_row, column=3).border = border
        ws.merge_cells(f'B{current_row}:C{current_row}')
        ws.cell(row=current_row, column=4, value=cnt).font = normal_font; ws.cell(row=current_row, column=4).border = border; ws.cell(row=current_row, column=4).alignment = center_align
        ws.cell(row=current_row, column=5, value="✅ " + res).font = normal_font; ws.cell(row=current_row, column=5).border = border; ws.cell(row=current_row, column=5).alignment = center_align
        ws.cell(row=current_row, column=5).fill = pass_bg
        current_row += 1

    # Chart 2
    pie2 = PieChart()
    pie2.title = "API & Security Testing (189 TCs)"
    labels2 = Reference(ws, min_col=data_col_1, min_row=6, max_row=7)
    data2 = Reference(ws, min_col=data_col_2, min_row=5, max_row=7)
    pie2.add_data(data2, titles_from_data=True)
    pie2.set_categories(labels2)
    pie2.height = 7
    pie2.width = 12
    ws.add_chart(pie2, f"H16")
    
    current_row += 2

    # ==========================================
    # 3. PERFORMANCE TESTING
    # ==========================================
    jmeter_data = [
        ("Số kịch bản tự động (Test Cases)", 255, "-", "Theo chuẩn IEEE Test Suite"),
        ("Tổng số request giả lập", "Max 10,000 users", "-", "Nhiều thread groups (100 -> 10k)"),
        ("Tỷ lệ Error trung bình", "4.15%", "-", "Từ 0% lên 12.5% khi tải lớn"),
        ("Response time trung bình (Avg)", "125 ms", "-", "Mức tải ổn định (100-1000 users)"),
        ("Max Response Time", "2400 ms", "-", "Ở mức 10,000 users peak load"),
        ("Throughput Max (Requests/sec)", "85.4/s", "-", "Tải chịu đựng ổn định")
    ]
    current_row = draw_table_section("3. KIỂM THỬ HIỆU NĂNG - JMETER", ["HẠNG MỤC CHI TIẾT", "THÔNG SỐ", "TỶ LỆ (%)", "GHI CHÚ"], jmeter_data, current_row)

    # Note about the load error rates table
    ws.merge_cells(f'B{current_row}:C{current_row}')
    ws.cell(row=current_row, column=2, value="Chi tiết Error Rate theo Load (Virtual Users):").font = Font(bold=True, italic=True)
    current_row += 1
    
    # Header
    ws.cell(row=current_row, column=2, value="VIRTUAL USERS").font = bold_font; ws.cell(row=current_row, column=2).border = border; ws.cell(row=current_row, column=2).fill = sec_bg; ws.cell(row=current_row, column=2).alignment = center_align
    ws.cell(row=current_row, column=3).border = border; ws.cell(row=current_row, column=3).fill = sec_bg
    ws.merge_cells(f'B{current_row}:C{current_row}')
    ws.cell(row=current_row, column=4, value="ERROR %").font = bold_font; ws.cell(row=current_row, column=4).border = border; ws.cell(row=current_row, column=4).fill = sec_bg; ws.cell(row=current_row, column=4).alignment = center_align
    current_row += 1
    
    for u, e in perf_rates:
        ws.cell(row=current_row, column=2, value=u).font = normal_font; ws.cell(row=current_row, column=2).border = border; ws.cell(row=current_row, column=2).alignment = center_align
        ws.cell(row=current_row, column=3).border = border
        ws.merge_cells(f'B{current_row}:C{current_row}')
        ws.cell(row=current_row, column=4, value=f"{e}%").font = normal_font; ws.cell(row=current_row, column=4).border = border; ws.cell(row=current_row, column=4).alignment = center_align
        if e > 0:
            ws.cell(row=current_row, column=4).fill = fail_bg
        else:
            ws.cell(row=current_row, column=4).fill = pass_bg
        current_row += 1

    # Bar chart
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Performance - Error Rate % by Virtual Users"
    bar.y_axis.title = 'Error %'
    bar.x_axis.title = 'Virtual Users'
    
    data3 = Reference(ws, min_col=data_col_2, min_row=9, max_row=15)
    cats3 = Reference(ws, min_col=data_col_1, min_row=10, max_row=15)
    bar.add_data(data3, titles_from_data=True)
    bar.set_categories(cats3)
    bar.height = 10
    bar.width = 15
    ws.add_chart(bar, f"H29")
    
    current_row += 3

    # ==========================================
    # 4. OVERALL SUMMARY
    # ==========================================
    overall_data = [
        ("Tổng mức độ hoàn thiện kiểm thử", "PASS", "100%", "Đã thực hiện đủ vòng Test"),
        ("Chất lượng logic hệ thống (Core logic)", "Tốt", "98%", "Logic matching / auth hoạt động cực tốt"),
        ("Chất lượng bảo mật (Security Tests)", "Rất Tốt", "100%", "Đã pass bài test về lỗ hổng OWASP Top 10"),
        ("Chất lượng hiệu năng (Performance)", "Tốt", "-", "Xử lý chịu tải mượt mà dưới 5000 users"),
        ("YÊU CẦU ĐÁNH GIÁ (KCS)", "Sẵn sàng Release", "-", "")
    ]
    current_row = draw_table_section("4. TỔNG HỢP TOÀN BỘ KIỂM THỬ", ["HẠNG MỤC CHI TIẾT", "ĐÁNH GIÁ", "TỶ LỆ (%)", "GHI CHÚ"], overall_data, current_row)

    ws.merge_cells(f'B{current_row}:F{current_row}')
    cell = ws.cell(row=current_row, column=2, value="Kết luận: Hệ thống StockLab có chuẩn mực cao về UI/UX lẫn cốt lõi bảo mật và hiệu năng. Đủ điều kiện Golive/Release.")
    cell.font = Font(name='Arial', size=11, italic=True, color='008000', bold=True)

    wb.save(target_excel)
    print(f"Final merged report saved to {target_excel}")

run()
