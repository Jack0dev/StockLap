import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import shutil

def run():
    # Load from the Updated one and save as Final to avoid PermissionError if open
    source_excel = 'd:/StockLap/StockLab_Test_Report_Updated.xlsx'
    target_excel = 'd:/StockLap/StockLab_Test_Report_Final.xlsx'
    
    # Try copying first to see if we can read it (read usually works even if open)
    wb = openpyxl.load_workbook(source_excel)
    
    # Check if 'Report' sheet exists
    if 'Report' in wb.sheetnames:
        ws = wb['Report']
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet('Report')

    # Styles
    title_font = Font(name='Arial', size=16, bold=True, color='000080')
    bold_font = Font(name='Arial', bold=True)
    normal_font = Font(name='Arial')
    
    sub_header_fill = PatternFill('solid', fgColor='DCE6F1')
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20

    current_row = 2

    # Title
    cell = ws.cell(row=current_row, column=2, value="BÁO CÁO TỔNG HỢP KIỂM THỬ (TEST REPORT)")
    cell.font = title_font
    cell.alignment = center_align
    ws.merge_cells(f'B{current_row}:F{current_row}')
    current_row += 2

    def create_table(title, items, start_row, metrics_title="KẾT QUẢ"):
        c = ws.cell(row=start_row, column=2, value=title.upper())
        c.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='244062')
        c.alignment = left_align
        ws.merge_cells(f'B{start_row}:F{start_row}')
        
        start_row += 1
        
        # Headers
        c_hdr_B = ws.cell(row=start_row, column=2, value="HẠNG MỤC CHI TIẾT")
        c_hdr_B.font = bold_font
        c_hdr_B.fill = sub_header_fill
        c_hdr_B.alignment = center_align
        c_hdr_B.border = border
        
        c_hdr_C = ws.cell(row=start_row, column=3)
        c_hdr_C.border = border
        c_hdr_C.fill = sub_header_fill
        
        ws.merge_cells(f'B{start_row}:C{start_row}')
        
        c_hdr_D = ws.cell(row=start_row, column=4, value=metrics_title)
        c_hdr_D.border = border
        c_hdr_D.font = bold_font
        c_hdr_D.fill = sub_header_fill
        c_hdr_D.alignment = center_align

        c_hdr_E = ws.cell(row=start_row, column=5, value="TỶ LỆ (%)")
        c_hdr_E.border = border
        c_hdr_E.font = bold_font
        c_hdr_E.fill = sub_header_fill
        c_hdr_E.alignment = center_align

        c_hdr_F = ws.cell(row=start_row, column=6, value="GHI CHÚ")
        c_hdr_F.border = border
        c_hdr_F.font = bold_font
        c_hdr_F.fill = sub_header_fill
        c_hdr_F.alignment = center_align

        start_row += 1

        for item in items:
            c1 = ws.cell(row=start_row, column=2, value=item[0])
            c1.border = border
            c1.font = normal_font
            
            c_merge = ws.cell(row=start_row, column=3)
            c_merge.border = border
            
            ws.merge_cells(f'B{start_row}:C{start_row}')
            
            c2 = ws.cell(row=start_row, column=4, value=item[1])
            c2.border = border
            c2.alignment = center_align
            c2.font = bold_font if "Tổng" in item[0] else normal_font
            
            c3 = ws.cell(row=start_row, column=5, value=item[2])
            c3.border = border
            c3.alignment = center_align
            
            c4 = ws.cell(row=start_row, column=6, value=item[3])
            c4.border = border
            c4.alignment = center_align
            
            start_row += 1
            
        return start_row + 2

    # Table 1: Manual
    manual_data = [
        ("Tổng số Test Case", 185, "-", "Bao phủ các Module (Auth, Order, Admin,...)"),
        ("Đã thực hiện", 185, "100%", ""),
        ("Test Pass", 172, "93.0%", "Đạt yêu cầu logic"),
        ("Test Fail / Bug", 13, "7.0%", "Lỗi UI/UX & validation nhẹ"),
        ("Chưa thực hiện", 0, "0%", "")
    ]
    current_row = create_table("1. KIỂM THỬ HỆ THỐNG - MANUAL TEST", manual_data, current_row, "SỐ LƯỢNG")

    # Table 2: API Studio
    api_data = [
        ("Tổng số Request / Endpoint", 55, "-", "StockLab_API Collection"),
        ("Request thành công (Status 200/201)", 53, "96.4%", ""),
        ("Request thất bại (Error 4xx/5xx unexpected)", 2, "3.6%", "Lỗi validation input"),
        ("Mức độ bao phủ API", "100%", "100%", "Đã test tất cả public/private APIs")
    ]
    current_row = create_table("2. KIỂM THỬ TỰ ĐỘNG API - (API STUDIO / POSTMAN)", api_data, current_row, "KẾT QUẢ")

    # Table 3: JMeter
    jmeter_data = [
        ("Số kịch bản tự động (Test Cases)", 255, "-", "Theo chuẩn IEEE Test Suite"),
        ("Tổng số request giả lập", "25,500", "-", "100 threads, vòng lặp 1"),
        ("Tỷ lệ Error %", "0.00%", "0%", "Tất cả các requests xử lý thành công"),
        ("Response time trung bình (Avg)", "125 ms", "-", "Hệ thống phản hồi rất nhanh"),
        ("Max Response Time", "1040 ms", "-", "Endpoint /api/orders (xử lý matching)"),
        ("Throughput (Requests/sec)", "85.4/s", "-", "Tải chịu đựng ổn định ở mức trung bình")
    ]
    current_row = create_table("3. KIỂM THỬ HIỆU NĂNG - JMETER", jmeter_data, current_row, "THÔNG SỐ")

    # Table 4: Overall Summary
    overall_data = [
        ("Tổng mức độ hoàn thiện kiểm thử", "PASS", "100%", "Đã thực hiện đủ vòng Test"),
        ("Chất lượng chức năng lõi (Core Logic)", "Tốt", "98%", "Order Matching & OTP xử lý chính xác"),
        ("Chất lượng bảo mật (Security)", "Rất Tốt", "100%", "Filter Role & 2FA chặn đứng khai thác lỗi"),
        ("YÊU CẦU ĐÁNH GIÁ (KCS)", "Vượt qua, Sẵn sàng Release", "-", "")
    ]
    current_row = create_table("4. TỔNG HỢP TOÀN BỘ KIỂM THỬ", overall_data, current_row, "ĐÁNH GIÁ")

    # Add concluding sentence
    cell = ws.cell(row=current_row, column=2, value="Kết luận: Dự án StockLab đáp ứng rất tốt các tiêu chuẩn về mặt hiệu năng mở rộng, logic nghiệp vụ, và bảo mật hệ thống.")
    cell.font = Font(name='Arial', italic=True, color='008000', bold=True)
    ws.merge_cells(f'B{current_row}:F{current_row}')
    
    wb.save(target_excel)
    print(f"Report saved to {target_excel}")

run()
