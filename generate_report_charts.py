import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference, Series
import shutil

def run():
    source_excel = 'd:/StockLap/StockLab_Test_Report_Final.xlsx'
    target_excel = 'd:/StockLap/StockLab_Test_Report_v2.xlsx'
    
    # Try copying to avoid permission errors
    try:
        shutil.copy(source_excel, target_excel)
    except Exception as e:
        print("Failed to copy", e)
        # Fallback to older source
        shutil.copy('d:/StockLap/StockLab_Test_Report_Updated.xlsx', target_excel)

    wb = openpyxl.load_workbook(target_excel)
    
    # Check if 'Report' sheet exists
    if 'Report' in wb.sheetnames:
        ws = wb['Report']
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet('Report')

    # Styles
    title_font = Font(name='Arial', size=16, bold=True, color='000080')
    bold_font = Font(name='Arial', bold=True)
    normal_font = Font(name='Arial')
    
    sub_header_fill = PatternFill('solid', fgColor='DCE6F1')
    pass_fill = PatternFill('solid', fgColor='C4D79B')
    fail_fill = PatternFill('solid', fgColor='E6B8B7')
    security_fill = PatternFill('solid', fgColor='002060')
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    data_col_1 = 27 # AA
    data_col_2 = 28 # AB
    
    # Write Chart Data
    # 1. Manual
    ws.cell(row=1, column=data_col_1, value="Manual")
    ws.cell(row=2, column=data_col_1, value="Pass")
    ws.cell(row=2, column=data_col_2, value=87)
    ws.cell(row=3, column=data_col_1, value="Fail")
    ws.cell(row=3, column=data_col_2, value=11)
    
    # 2. API
    ws.cell(row=5, column=data_col_1, value="API")
    ws.cell(row=6, column=data_col_1, value="Pass")
    ws.cell(row=6, column=data_col_2, value=292)
    ws.cell(row=7, column=data_col_1, value="Fail")
    ws.cell(row=7, column=data_col_2, value=5)
    
    # 3. Perf
    ws.cell(row=9, column=data_col_1, value="Users")
    ws.cell(row=9, column=data_col_2, value="Error %")
    perf_data = [
        ("100 users", 0),
        ("500 users", 5),
        ("1000 users", 11),
        ("2000 users", 11),
        ("5000 users", 50),
        ("10000 users", 73)
    ]
    for i, (u, e) in enumerate(perf_data):
        ws.cell(row=10+i, column=data_col_1, value=u)
        ws.cell(row=10+i, column=data_col_2, value=e)

    current_row = 2

    def write_cell(r, c, val, f=normal_font, bg=None, align=center_align):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = border
        cell.font = f
        cell.alignment = align
        if bg:
            cell.fill = bg
        return cell

    # ==========================
    # 1. MANUAL TESTING
    # ==========================
    ws.merge_cells(f'B{current_row}:D{current_row}')
    c = ws.cell(row=current_row, column=2, value="[ MANUAL ]")
    c.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='000080')
    c.alignment = center_align
    current_row += 1
    
    write_cell(current_row, 2, "KIỂM TRA", bold_font, PatternFill('solid', fgColor='002060')).font = Font(color='FFFFFF', bold=True)
    write_cell(current_row, 3, "SỐ LƯỢNG", bold_font, PatternFill('solid', fgColor='002060')).font = Font(color='FFFFFF', bold=True)
    write_cell(current_row, 4, "PHẦN TRĂM", bold_font, PatternFill('solid', fgColor='002060')).font = Font(color='FFFFFF', bold=True)
    current_row += 1
    
    manual_rows = [
        ("TEST CASE", 98, "100%", None),
        ("TEST THỰC HIỆN", 98, "100%", None),
        ("TEST CHƯA THỰC HIỆN", 0, "0%", None),
        ("TEST FAIL", 11, "11.22%", fail_fill),
        ("TEST PASS", 87, "88.78%", pass_fill),
    ]
    for name, count, pct, bg in manual_rows:
        write_cell(current_row, 2, name, normal_font, bg)
        write_cell(current_row, 3, count, normal_font)
        write_cell(current_row, 4, pct, normal_font)
        current_row += 1
        
    pie1 = PieChart()
    pie1.title = "Manual Testing (98 TC)"
    labels1 = Reference(ws, min_col=data_col_1, min_row=2, max_row=3)
    data1 = Reference(ws, min_col=data_col_2, min_row=2, max_row=3)
    pie1.add_data(data1, titles_from_data=False)
    pie1.set_categories(labels1)
    ws.add_chart(pie1, f"F2")
    
    current_row += 3
    
    # ==========================
    # 2. API & SECURITY TESTING
    # ==========================
    ws.merge_cells(f'B{current_row}:D{current_row}')
    c = ws.cell(row=current_row, column=2, value="[ API & SECURITY ]")
    c.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='000080')
    c.alignment = center_align
    current_row += 1
    
    write_cell(current_row, 2, "KIỂM TRA", bold_font, PatternFill('solid', fgColor='002060')).font = Font(color='FFFFFF', bold=True)
    write_cell(current_row, 3, "SỐ LƯỢNG", bold_font, PatternFill('solid', fgColor='002060')).font = Font(color='FFFFFF', bold=True)
    write_cell(current_row, 4, "PHẦN TRĂM", bold_font, PatternFill('solid', fgColor='002060')).font = Font(color='FFFFFF', bold=True)
    current_row += 1
    
    api_rows = [
        ("TỔNG TEST CASES", 297, "100%", None),
        ("FUNCTIONAL TESTS", "", "54.88%", None),
        ("SECURITY TESTS", 134, "45.12%", None),
        ("TEST PASS", 292, "98.32%", pass_fill),
        ("TEST FAIL", 5, "1.68%", fail_fill),
    ]
    for name, count, pct, bg in api_rows:
        write_cell(current_row, 2, name, normal_font, bg)
        write_cell(current_row, 3, count, normal_font)
        write_cell(current_row, 4, pct, normal_font)
        current_row += 1
        
    ws.merge_cells(f'B{current_row}:D{current_row}')
    # Note: openpyxl border works best on left-most cell of merge, but sometimes we need to border all cells
    c_sec = write_cell(current_row, 2, "Chi tiết Security:", bold_font, None, left_align)
    ws.cell(row=current_row, column=3).border = border
    ws.cell(row=current_row, column=4).border = border
    current_row += 1
    
    write_cell(current_row, 2, "LOẠI TẤN CÔNG", bold_font, PatternFill('solid', fgColor='002060')).font = Font(color='FFFFFF', bold=True)
    write_cell(current_row, 3, "SỐ TEST", bold_font, PatternFill('solid', fgColor='002060')).font = Font(color='FFFFFF', bold=True)
    write_cell(current_row, 4, "KẾT QUẢ", bold_font, PatternFill('solid', fgColor='002060')).font = Font(color='FFFFFF', bold=True)
    current_row += 1
    
    sec_rows = [
        ("SQL Injection", 38, "✅ Chặn 100%"),
        ("XSS Attack", 38, "✅ Chặn 100%"),
        ("Path Traversal", 46, "✅ Chặn 100%"),
        ("Unauthorized Access", 12, "✅ Chặn 100%"),
    ]
    for name, count, pct in sec_rows:
        write_cell(current_row, 2, name, normal_font, PatternFill('solid', fgColor='F2DCDB'))
        write_cell(current_row, 3, count, bold_font)
        write_cell(current_row, 4, pct, normal_font)
        current_row += 1

    pie2 = PieChart()
    pie2.title = "Katalon API Testing (297 TC)"
    labels2 = Reference(ws, min_col=data_col_1, min_row=6, max_row=7)
    data2 = Reference(ws, min_col=data_col_2, min_row=6, max_row=7)
    pie2.add_data(data2, titles_from_data=False)
    pie2.set_categories(labels2)
    ws.add_chart(pie2, f"F15")
    
    current_row += 3

    # ==========================
    # 3. PERFORMANCE TESTING
    # ==========================
    ws.merge_cells(f'B{current_row}:D{current_row}')
    c = ws.cell(row=current_row, column=2, value="[ PERFORMANCE TESTING ]")
    c.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='000080')
    c.alignment = center_align
    current_row += 1
    
    write_cell(current_row, 2, "CHỈ TIÊU", bold_font, PatternFill('solid', fgColor='002060')).font = Font(color='FFFFFF', bold=True)
    write_cell(current_row, 3, "USERS", bold_font, PatternFill('solid', fgColor='002060')).font = Font(color='FFFFFF', bold=True)
    write_cell(current_row, 4, "ERROR %", bold_font, PatternFill('solid', fgColor='002060')).font = Font(color='FFFFFF', bold=True)
    current_row += 1
    
    for u, e in perf_data:
        write_cell(current_row, 2, "Error Rate By Load", normal_font)
        write_cell(current_row, 3, u, normal_font)
        write_cell(current_row, 4, str(e) + "%", normal_font, fail_fill if e > 0 else pass_fill)
        current_row += 1

    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Performance - Error Rate by Load"
    bar.y_axis.title = 'Error %'
    bar.x_axis.title = 'Virtual Users'
    
    data3 = Reference(ws, min_col=data_col_2, min_row=9, max_row=15)
    cats3 = Reference(ws, min_col=data_col_1, min_row=10, max_row=15)
    bar.add_data(data3, titles_from_data=True)
    bar.set_categories(cats3)
    ws.add_chart(bar, f"F28")
    
    current_row += 3
    
    wb.save(target_excel)
    print(f"Detailed Report with Charts saved to {target_excel}")

run()
